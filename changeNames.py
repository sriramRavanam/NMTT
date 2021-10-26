### python script to change the old names to the official district names.
import glob
import openpyxl as op

namesToChangeFile = "./result/NamesToChange"

# lines = []
with open(namesToChangeFile,"r") as file:
    lines = file.readlines()


lines = list(map(lambda x: x.strip() , lines))

officialNamesFile = "./result/officialNamesKarnataka.txt"

# official names
with open(officialNamesFile,"r") as file:
    officialNames = file.readlines()

officialNames = list(map(lambda x: x.strip(), officialNames))
officialNamesDict = {}

for i in range(len(officialNames)):
    officialNamesDict[i] = officialNames[i]


dictValues = {}

for s in lines:
    split = s.split(" - ")
    value = split[-1]
    if value in officialNames:
        value = officialNames.index(value) + 1
    value = int(value)
    key = "".join(split[:-1])

    dictValues[key] = officialNamesDict[value-1]


# excelFilesPath = "test2"
excelFilesPath = "mathTransformed"
typeOfFile = "*.xlsx"

excelFilesList = glob.glob(excelFilesPath+"/"+typeOfFile)

for excel in excelFilesList:
    # we take the path to excel files

    wb = op.load_workbook(excel)
    ws = wb.worksheets[0]

    maxRows = ws.max_row + 1
    
    for rowIdx in range(3,maxRows):
        district = ws.cell(row = rowIdx, column = 7).value
        
        if district in dictValues:
            ws.cell(row = rowIdx, column= 7).value = dictValues[district]

    wb.save(excel)




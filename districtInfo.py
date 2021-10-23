### script to get the count of districts and the number of people per district trained.
import openpyxl as op
import glob

location = "mathTransformed"
# location = "test2"

destination = "result"


Districts = {}
# countPerDistrict = {}

# districts are in G
pattern = "*.xlsx"

files = glob.glob(location+"/"+pattern)

for file in files:
    # here file is the path to the excel workbook
    
    wb = op.load_workbook(filename=file)
    ws = wb.worksheets[0]

    maxRows = ws.max_row + 1

    for i in range(3,maxRows):
        districtName = ws.cell(row=i,column=7).value

        if districtName not in Districts.keys():
            Districts[districtName] = 1
        else :
            Districts[districtName] = Districts[districtName] + 1


keys  = list(Districts.keys())

keys.sort()
# print(keys)
# # print(Districts)

with open(destination+"/districtCount.txt","w") as file:
    for s in keys :
        file.write(s+" - "+str(Districts[s])+" \n")

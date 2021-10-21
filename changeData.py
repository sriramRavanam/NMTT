import os
import openpyxl as op
from copy import copy
import glob 


dir = "mathExcels"
destination = "mathTransformed"
pattern = "*.xlsx"
# excelFiles = os.listdir(dir+"\.") // didnt allow use of regular expressions, hence had to use glob
excelFiles = glob.glob(dir+"/"+pattern)

for file in excelFiles:
    filePath = file
    
    fileName = file.split("\\")[-1]

    destinationPath = destination+"/"+fileName
    #old workbook
    wb = op.load_workbook(filename=filePath)
    oldSheet = wb.worksheets[0]
    #new workbook
    newWb = op.Workbook()
    newSheet = newWb.active

    ### below code copies the headings for the first two rows and also adds additional fields for the data.

    newSheet["A1"].value = oldSheet["A1"].value
    newSheet["B1"].value = oldSheet["B1"].value
    newSheet["C1"].value = oldSheet["C1"].value
    newSheet["D1"].value = oldSheet["D1"].value
    
    newSheet["A2"].value = oldSheet["A2"].value
    newSheet["B2"].value = oldSheet["B2"].value
    newSheet["C2"].value = oldSheet["C2"].value
    newSheet["D2"].value = oldSheet["D2"].value

    newSheet["E2"].value = "Name"
    newSheet["F2"].value = "District"


    # newSheet["E2"].value = "Name"
    # newSheet["F2"].value = "Designation"
    # newSheet["G2"].value = "School-Name"
    # newSheet["H2"].value = "Town"
    # newSheet["I2"].value = "District"

    #code works till this part
    ### the for loop goes through the data and splits it into required format.

    maxRows = oldSheet.max_row
    maxCols = oldSheet.max_column

    #cuz table row data starts from row 3
    for i in range(3,maxRows+1):
        for j in range(1,maxCols+1):
            cell = oldSheet.cell(row=i,column=j)
            if(j==2):
                nameAndAddress = cell.value
                newSheet.cell(row=i,column=j).value = nameAndAddress
                continue
            newSheet.cell(row=i,column=j).value = cell.value
        # nameAndAddress = oldSheet.cell(row = i, column = 2).value

        # print(nameAndAddress)
        value = nameAndAddress.split("\n")
        fullName = value[0]
        Address = value[-1].split(", ")
        district = Address[-1]
        district = district[:-1]

        newSheet.cell(row=i,column=5).value = fullName
        # print(district)
        newSheet.cell(row=i,column=6).value = district
        
            ###Approach 2
        # print(Address)
        # address = Address.split(", ")
        # address[-1] = address[-1][:-1] #to remove the last '.'

        # newSheet.cell(row=i,column=5).value = fullName
        
        # try:
        #     newSheet.cell(row=i,column=6).value = address[-4]
        # except:
        #     newSheet.cell(row=i,column=6).value = None

        # newSheet.cell(row=i,column=7).value = address[-3]
        # newSheet.cell(row=i,column=8).value = address[-2]
        # newSheet.cell(row=i,column=9).value = address[-1]

        ###Approach 1 
    # for idx,row in enumerate(oldSheet.iter_rows(min_row=3)):
    #     ### old data has the following format
    #     ### SL.No | Name,Addr | Before Score | After Score
        
    #     slno = "A"+str(idx)
    #     nameAddr = "B"+str(idx)
    #     bfScr = "C"+str(idx)
    #     afScr = "D"+str(idx)
    #     name = "E"+str(idx)
    #     schNam = "F"+str(idx)
    #     distr = "H"+str(idx)
    #     townNam = "G"+str(idx)

    #     newSheet[slno].value = oldSheet[slno].value
    #     newSheet[nameAddr].value = oldSheet[nameAddr].value
    #     newSheet[bfScr].value = oldSheet[bfScr].value
    #     newSheet[afScr].value = oldSheet[afScr].value

    #     nameAddress = oldSheet[nameAddr].value
    #     fullName , Address = nameAddress.split("\n")
    #     schoolName,town,district = Address.split(", ")
    #     district = district[:-1] #to remove the last '.'


    #     newSheet[name].value = fullName
    #     newSheet[schNam].value = schoolName
    #     newSheet[distr].value = district
    #     newSheet[townNam].value = town


    newWb.save(filename=destinationPath)




    # val = ws["B3"].value
    # val = val.split("\n") #\n separates the name and the address
    # name,address = val 
    # schoolName,town,district = address.split(", ")  # 
    # district = district[:-1]    #everything except last character because '.' is present at the end.

    # the 3rd column typically has NAME\nADDRESS.
    # 
    # address has schoolName,town,district

    # print(schoolName,district,name,town)


    # first add a district column to the sheet which is on the E line
    # cellD2 = ws["D2"]
    # fontD2 = cellD2.font
    # cellE2 = ws["E2"]
    # if cellD2.has_style:
    #         cellE2.font = copy(cellD2.font)
    #         cellE2.border = copy(cellD2.border)
    #         cellE2.fill = copy(cellD2.fill)
    #         cellE2.number_format = copy(cellD2.number_format)
    #         cellE2.protection = copy(cellD2.protection)
    #         cellE2.alignment = copy(cellD2.alignment)

    # cellE2.value = "District"

    # for row in ws.iter_rows(min_row=2,values_only=False):  # values_only was true
        # for value in row:
        #     print(value)

        # print(row[2].value)
        # row[5].value = 5
        # print(row[5].value)

        # rowNum = row[0].row    # 1 based not 0
        

    